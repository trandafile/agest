"""Export XLSX di rendicontazione (formato Amendola)."""

from __future__ import annotations

import io

import openpyxl

from src.lib.rendiconto_xlsx import build_rendiconto_xlsx, nome_file


def _genera(**kw) -> openpyxl.Workbook:
    dati = dict(
        anno=2026,
        mese=2,
        cognome="Rossi",
        nome="Mario",
        codice_fiscale="RSSMRA80A01H501U",
        cup="H23C22001000003",
        soggetto_attuatore="ANTECNICA SRLS",
        titolo_progetto="Progetto Demo",
        tipo_progetto="Ricerca Industriale e Sviluppo Sperimentale",
        monte_ore_annuo=1720,
        ore_ri={2: 5, 3: 3},
        ore_ss={2: 2},
        ore_altri_progetti={4: 8},
    )
    dati.update(kw)
    return openpyxl.load_workbook(io.BytesIO(build_rendiconto_xlsx(**dati)))


def test_struttura_base():
    wb = _genera()
    ws = wb.active
    # troncato a 31 char (limite Excel), come nel modello originale
    assert ws.title == "periodo_2026-02_progH23C2200100"
    assert ws["A3"].value.startswith("TIMESHEET PER RENDICONTAZIONE")
    assert ws["A7"].value == "CUP del progetto:"
    assert ws["B7"].value == "H23C22001000003"
    # griglia: intestazione giorni a riga 12 (febbraio 2026 = 28 giorni)
    assert ws.cell(12, 2).value == 1
    assert ws.cell(12, 29).value == 28
    assert ws.cell(12, 30).value == "Tot"


def test_valori_e_totali():
    wb = _genera()
    ws = wb.active
    # riga RI (13): giorno 2 -> col 3
    assert ws.cell(13, 3).value == 5
    assert ws.cell(13, 30).value == 8  # 5+3
    assert ws.cell(14, 30).value == 2  # SS
    assert ws.cell(15, 30).value == 8  # altri progetti
    assert ws.cell(16, 30).value == 0  # didattica vuota
    # Totale (riga 18): giorno 2 = 5+2
    assert ws.cell(18, 1).value == "Totale"
    assert ws.cell(18, 3).value == 7
    assert ws.cell(18, 30).value == 18


def test_firme_presenti():
    ws = _genera().active
    testi = [c.value for row in ws.iter_rows(min_row=19, max_row=23) for c in row]
    assert "Firmato da (Nome e Cognome):" in testi
    assert "Firma:" in testi


def test_mese_31_giorni_e_nome_file():
    ws = _genera(mese=3).active
    assert ws.cell(12, 32).value == 31
    assert ws.cell(12, 33).value == "Tot"
    assert nome_file("Rossi", 2026, 3) == "Rossi2026_03.xlsx"


def test_logo_inserito():
    # PNG 1x1 valido
    png = (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
        b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\xcf\xc0"
        b"\x00\x00\x00\x03\x00\x01\x9a\x92\xdeA\x00\x00\x00\x00IEND\xaeB`\x82"
    )
    ws = _genera(logo=png).active
    assert len(ws._images) == 1
