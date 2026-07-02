"""Parser + export del Libro Cassa ANTECNICA."""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

import openpyxl

from src.lib.export_contabile import build_libro_cassa_xlsx
from src.lib.import_contabile import (
    leggi_workbook,
    parse_libro_cassa,
    parse_progetto,
    saldo,
)

D = Decimal


def test_parse_libro_cassa_salta_righe_riepilogo():
    righe = [
        [None, None, None, "SALDO", 100, None, None, None, None],
        [None, None, None, "Tot ingressi", 200, None, None, None, None],
        [None, None, None, "Tot Uscite", -100, None, None, None, None],
        [
            "Data",
            "Descrizione Transazione",
            "N. Fattura",
            "Tipo Transazione",
            "Importo (€)",
            "Categoria",
            "Progetto",
            "Persona/Contatto",
            "Note",
        ],
        [
            date(2026, 1, 2),
            "Bonifico cliente",
            None,
            "Entrata",
            1000,
            "Vendita",
            "RUSC",
            None,
            None,
        ],
        [
            date(2026, 1, 3),
            "Acquisto HW",
            "FT-1",
            "Uscita",
            -250.5,
            "Hardware",
            None,
            "Fornitore",
            "urgente",
        ],
        [None, None, None, None, None, None, None, None, None],
    ]
    movs = parse_libro_cassa(righe)
    assert len(movs) == 2
    assert movs[0]["segno"] == "entrata" and movs[0]["importo"] == D("1000")
    assert movs[0]["progetto_label"] == "RUSC"
    assert movs[1]["segno"] == "uscita" and movs[1]["importo"] == D("250.5")
    assert movs[1]["n_fattura"] == "FT-1"
    assert saldo(movs) == D("749.5")


def test_parse_libro_cassa_scarta_date_invalide():
    righe = [
        [
            "Data",
            "Descrizione Transazione",
            "N. Fattura",
            "Tipo Transazione",
            "Importo (€)",
            "Categoria",
            "Progetto",
            "Persona/Contatto",
            "Note",
        ],
        ["25/5/0206", "Refuso anno", None, "Uscita", -41, "Hardware", None, None, None],
        [date(2026, 5, 25), "Buono", None, "Uscita", -10, "Hardware", None, None, None],
    ]
    movs = parse_libro_cassa(righe)
    assert len(movs) == 1  # la riga col 2026 valida; il refuso 0206 scartato


def test_parse_progetto_info_e_calendario():
    righe = [
        ["SEZIONE: INFO GENERALI", None],
        ["Acronimo", "RUSC"],
        ["Nome esteso", "Progetto RUSC esteso"],
        ["Ente finanziatore", "PNRR"],
        ["Costo complessivo", 350712.5],
        ["Finanziamento complessivo", 283350.2],
        ["Tot spese personale", 130712.5],
        ["Tot acquisti", 220000.0],
        ["Data inizio", date(2024, 1, 1)],
        ["Data fine", date(2026, 12, 31)],
        [None, None],
        ["SEZIONE: CALENDARIO MOVIMENTI", None],
        ["Descrizione", "Ingresso/Uscita", "Importo (€)", "Data attesa", "Completata"],
        ["Anticipo", "Ingresso", 100000, date(2026, 3, 1), True],
        ["Saldo fornitore", "Uscita", 20000, date(2026, 6, 1), None],
    ]
    p = parse_progetto(righe)
    assert p["info"]["acronimo"] == "RUSC"
    assert p["info"]["ente_finanziatore"] == "PNRR"
    assert p["info"]["costo_complessivo"] == D("350712.5")
    assert p["info"]["data_fine"] == date(2026, 12, 31)
    assert len(p["calendario"]) == 2
    assert p["calendario"][0]["segno"] == "entrata"
    assert p["calendario"][0]["completata"] is True
    assert p["calendario"][1]["segno"] == "uscita"


def test_export_import_roundtrip():
    movimenti = {
        2026: [
            {
                "data": date(2026, 1, 2),
                "importo": D("1000"),
                "segno": "entrata",
                "descrizione": "Cliente",
                "n_fattura": "F1",
                "categoria": "Vendita",
                "persona_contatto": None,
                "note": None,
                "iniziativa_id": "X",
                "progetto_label": "RUSC",
            },
            {
                "data": date(2026, 1, 3),
                "importo": D("300"),
                "segno": "uscita",
                "descrizione": "HW",
                "n_fattura": None,
                "categoria": "Hardware",
                "persona_contatto": "Forn",
                "note": None,
                "iniziativa_id": None,
                "progetto_label": None,
            },
        ]
    }
    xlsx = build_libro_cassa_xlsx(movimenti, {"X": "RUSC"})
    # rileggo con openpyxl e verifico header + saldo
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert "Libro Cassa 2026" in wb.sheetnames
    ws = wb["Libro Cassa 2026"]
    assert ws["D1"].value == "SALDO" and ws["E1"].value == 700.0
    assert ws.cell(4, 1).value == "Data"
    # roundtrip: reimporto e ottengo gli stessi 2 movimenti
    sez = leggi_workbook(xlsx)
    movs = sez["libri"][2026]
    assert len(movs) == 2 and saldo(movs) == D("700")
    assert movs[0]["progetto_label"] == "RUSC"
