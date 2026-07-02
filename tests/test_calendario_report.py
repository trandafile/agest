"""Calendario (ICS, griglia) e report attività (Markdown/XLSX)."""

from __future__ import annotations

import io
from datetime import date
from types import SimpleNamespace

import openpyxl

from src.lib.calendario import (
    build_ics,
    eventi_per_giorno,
    link_google_calendar,
    settimane_mese,
)
from src.lib.report_attivita import report_markdown, tasks_xlsx


def test_settimane_mese_griglia():
    sm = settimane_mese(2026, 2)  # febbraio 2026
    assert all(len(w) == 7 for w in sm)
    giorni = [d for w in sm for d in w if d is not None]
    assert len(giorni) == 28
    assert giorni[0] == date(2026, 2, 1) and giorni[-1] == date(2026, 2, 28)


def test_build_ics():
    ev = [
        {
            "data": date(2026, 7, 2),
            "titolo": "Consegna",
            "tipo": "deliverable",
            "progetto": "RUSC",
            "pagamento": False,
        },
        {
            "data": date(2026, 7, 15),
            "titolo": "Acconto",
            "tipo": "milestone",
            "progetto": "IOT",
            "pagamento": True,
        },
    ]
    ics = build_ics(ev)
    assert ics.startswith("BEGIN:VCALENDAR")
    assert ics.count("BEGIN:VEVENT") == 2
    assert "DTSTART;VALUE=DATE:20260702" in ics
    assert "SUMMARY:📦 Consegna [RUSC]" in ics


def test_link_google_calendar():
    u = link_google_calendar("Task X", date(2026, 9, 15), "note")
    assert u.startswith("https://calendar.google.com/calendar/render")
    assert "action=TEMPLATE" in u
    assert "20260915%2F20260916" in u  # DTSTART/DTEND (giorno successivo)


def test_eventi_per_giorno():
    ev = [
        {"data": date(2026, 7, 2), "titolo": "a"},
        {"data": date(2026, 7, 2), "titolo": "b"},
        {"data": date(2026, 7, 3), "titolo": "c"},
    ]
    m = eventi_per_giorno(ev)
    assert len(m[date(2026, 7, 2)]) == 2 and len(m[date(2026, 7, 3)]) == 1


def _task(**kw):
    base = dict(
        id=kw.get("id"),
        iniziativa_id=None,
        deliverable_id=None,
        parent_task_id=None,
        titolo="T",
        owner_id="o",
        supervisor_id=None,
        stato="da_fare",
        priorita="alta",
        ore_stimate=None,
        scadenza=None,
        completato_il=None,
    )
    base.update(kw)
    return SimpleNamespace(**base)


def test_report_markdown_albero():
    ini = SimpleNamespace(id="i1", titolo="Progetto Demo", acronimo="DEMO", codice=None)
    deliv = SimpleNamespace(
        id="d1", titolo="Report", tipo="report", scadenza=date(2026, 9, 30)
    )
    t = _task(
        id="t1",
        iniziativa_id="i1",
        deliverable_id="d1",
        titolo="Scrivi",
        scadenza=date(2026, 9, 15),
        ore_stimate=10,
    )
    sub = _task(
        id="s1",
        iniziativa_id="i1",
        deliverable_id="d1",
        parent_task_id="t1",
        titolo="Bozza",
    )
    libero = _task(id="t2", iniziativa_id="i1", titolo="Task libero")
    md = report_markdown(
        [ini],
        {"i1": [deliv]},
        [t, sub, libero],
        {"o": "Anna"},
        lambda i: i.acronimo,
    )
    assert "## 📁 DEMO" in md
    assert "### 📦 Report" in md
    assert "**Scrivi**" in md and "Bozza" in md
    assert "(senza deliverable)" in md and "Task libero" in md


def test_tasks_xlsx():
    t = _task(id="t1", titolo="Scrivi", stato="in_corso", ore_stimate=8)
    x = tasks_xlsx([t], {"o": "Anna"}, {})
    wb = openpyxl.load_workbook(io.BytesIO(x))
    ws = wb.active
    assert ws.cell(1, 1).value == "Progetto"
    assert ws.cell(2, 2).value == "Scrivi"
    assert ws.cell(2, 4).value == "Anna"
