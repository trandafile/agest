"""Export XLSX del timesheet per rendicontazione (formato "Amendola").

Replica il layout del modello usato per i SAL MIUR/PNRR:
- intestazione anagrafica + progetto (CUP, soggetto attuatore, monte ore),
- griglia giorni (1..N) x attivita' (RI - CUP, SS - CUP, Altri progetti,
  Attivita' didattica, Altro) con colonna Totale,
- riga Totale e blocco firme; logo del progetto in alto se disponibile.
"""

from __future__ import annotations

import calendar
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MESI_IT = (
    "Gennaio",
    "Febbraio",
    "Marzo",
    "Aprile",
    "Maggio",
    "Giugno",
    "Luglio",
    "Agosto",
    "Settembre",
    "Ottobre",
    "Novembre",
    "Dicembre",
)

_BORDO = Border(*(Side(style="thin"),) * 4)
_GRIGIO = PatternFill("solid", fgColor="EEEEEE")
_BOLD = Font(bold=True)


def build_rendiconto_xlsx(
    *,
    anno: int,
    mese: int,
    cognome: str,
    nome: str,
    codice_fiscale: str | None,
    cup: str | None,
    soggetto_attuatore: str,
    titolo_progetto: str,
    tipo_progetto: str | None,
    monte_ore_annuo: int | None,
    ore_ri: dict[int, int],
    ore_ss: dict[int, int],
    ore_altri_progetti: dict[int, int],
    logo: bytes | None = None,
) -> bytes:
    """Genera il file XLSX e ritorna i byte.

    `ore_*`: {giorno_del_mese: ore}. Le righe "Attivita' didattica" e "Altro"
    restano vuote (la didattica e' fuori scope per ANTECNICA, la riga esiste
    per fedelta' al modello).
    """
    n_giorni = calendar.monthrange(anno, mese)[1]
    col_tot = 1 + n_giorni + 1  # A=etichette, B..=giorni, ultima=Totale
    ultima = get_column_letter(col_tot)

    wb = Workbook()
    ws = wb.active
    cup_slug = (cup or "NOCUP").replace(" ", "")
    ws.title = f"periodo_{anno}-{mese:02d}_prog{cup_slug}"[:31]

    def merge_row(riga: int, da: int, a: int) -> None:
        ws.merge_cells(start_row=riga, start_column=da, end_row=riga, end_column=a)

    # --- Logo (righe 1-2 lasciate libere per l'immagine) ---------------------
    if logo:
        try:
            img = XLImage(io.BytesIO(logo))
            scala = min(1.0, 120 / max(img.height, 1))
            img.height = int(img.height * scala)
            img.width = int(img.width * scala)
            ws.add_image(img, "A1")
        except Exception:  # logo corrotto: si procede senza
            pass

    # --- Intestazione ---------------------------------------------------------
    merge_row(3, 1, col_tot)
    ws.cell(3, 1, "TIMESHEET PER RENDICONTAZIONE PERSONALE:").font = _BOLD

    q1 = col_tot // 4  # blocchi da un quarto riga per le coppie label/valore
    coppie_r4 = [("Anno:", anno), ("Mese:", MESI_IT[mese - 1])]
    coppie_r5 = [("Cognome:", cognome.upper()), ("Nome:", nome.upper())]
    for riga, coppie in ((4, coppie_r4), (5, coppie_r5)):
        blocchi = [
            (1, q1),
            (q1 + 1, 2 * q1),
            (2 * q1 + 1, 3 * q1),
            (3 * q1 + 1, col_tot),
        ]
        for i, (label, val) in enumerate(coppie):
            la, lb = blocchi[2 * i]
            va, vb = blocchi[2 * i + 1]
            merge_row(riga, la, lb)
            merge_row(riga, va, vb)
            ws.cell(riga, la, label).font = _BOLD
            ws.cell(riga, va, val)

    righe_info = [
        ("Codice Fiscale:", codice_fiscale or ""),
        ("CUP del progetto:", cup or ""),
        ("Soggetto attuatore:", soggetto_attuatore),
        ("Titolo del progetto:", titolo_progetto),
        ("Tipo del progetto:", tipo_progetto or ""),
        ("Monte ore lavorative annuo previsto:", monte_ore_annuo or ""),
    ]
    for i, (label, val) in enumerate(righe_info):
        riga = 6 + i
        ws.cell(riga, 1, label).font = _BOLD
        merge_row(riga, 2, col_tot)
        ws.cell(riga, 2, val)

    # --- Griglia ----------------------------------------------------------------
    r_head = 12
    ws.cell(r_head, 1, "Attività svolta sul Progetto\\Day").font = _BOLD
    ws.cell(r_head, 1).fill = _GRIGIO
    for g in range(1, n_giorni + 1):
        c = ws.cell(r_head, 1 + g, g)
        c.font = _BOLD
        c.fill = _GRIGIO
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(1 + g)].width = 4.2
    c = ws.cell(r_head, col_tot, "Tot")
    c.font = _BOLD
    c.fill = _GRIGIO
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions[ultima].width = 7

    etichetta_cup = f" - {cup}" if cup else ""
    righe_griglia = [
        (f"Ricerca Industriale{etichetta_cup}", ore_ri),
        (f"Sviluppo sperimentale{etichetta_cup}", ore_ss),
        ("Altri progetti", ore_altri_progetti),
        ("Attività didattica", {}),
        ("Altro", {}),
    ]
    for i, (label, valori) in enumerate(righe_griglia):
        riga = r_head + 1 + i
        ws.cell(riga, 1, label)
        tot = 0
        for g in range(1, n_giorni + 1):
            v = int(valori.get(g, 0) or 0)
            if v:
                ws.cell(riga, 1 + g, v)
                tot += v
        ws.cell(riga, col_tot, tot).font = _BOLD

    r_tot = r_head + 1 + len(righe_griglia)
    ws.cell(r_tot, 1, "Totale").font = _BOLD
    ws.cell(r_tot, 1).fill = _GRIGIO
    tot_generale = 0
    for g in range(1, n_giorni + 1):
        tot_g = sum(int(v.get(g, 0) or 0) for _, v in righe_griglia)
        cella = ws.cell(r_tot, 1 + g, tot_g)
        cella.font = _BOLD
        cella.fill = _GRIGIO
        tot_generale += tot_g
    c = ws.cell(r_tot, col_tot, tot_generale)
    c.font = _BOLD
    c.fill = _GRIGIO

    # bordi della griglia
    for riga in range(r_head, r_tot + 1):
        for col in range(1, col_tot + 1):
            ws.cell(riga, col).border = _BORDO

    # --- Firme -------------------------------------------------------------------
    meta = col_tot // 2
    r_firma = r_tot + 2
    for i, label in enumerate(("Firmato da (Nome e Cognome):", "Data:", "Firma:")):
        riga = r_firma + i
        merge_row(riga, 1, meta // 2)
        ws.cell(riga, 1, label).font = _BOLD
        merge_row(riga, meta + 1, meta + meta // 2)
        ws.cell(riga, meta + 1, label).font = _BOLD

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def nome_file(cognome: str, anno: int, mese: int) -> str:
    return f"{cognome.capitalize()}{anno}_{mese:02d}.xlsx"


def periodo_giorno(d: date) -> int:
    return d.day
