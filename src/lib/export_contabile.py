"""Export dei dati contabili nel formato «Libro Cassa» ANTECNICA (XLSX).

Riproduce il layout del Google Sheet (righe SALDO/Tot in alto, header, dati)
così che il file possa essere ricaricato o incollato nel foglio Google.
Un foglio «Libro Cassa <anno>» per anno + un foglio per progetto (INFO
GENERALI + CALENDARIO MOVIMENTI).
"""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_BOLD = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="CFE2F3")
_INTESTA = [
    "Data",
    "Descrizione Transazione",
    "N. Fattura",
    "Tipo Transazione",
    "Importo (€)",
    "Categoria",
    "Progetto",
    "Persona/Contatto",
    "Note",
]


def _saldo(movimenti: list[dict]) -> tuple[Decimal, Decimal, Decimal]:
    ent = sum(
        (Decimal(str(m["importo"])) for m in movimenti if m["segno"] == "entrata"),
        Decimal("0"),
    )
    usc = sum(
        (Decimal(str(m["importo"])) for m in movimenti if m["segno"] == "uscita"),
        Decimal("0"),
    )
    return ent - usc, ent, -usc


def _foglio_libro(wb: Workbook, anno, movimenti: list[dict], acronimi: dict) -> None:
    ws = wb.create_sheet(f"Libro Cassa {anno}" if anno else "Libro Cassa")
    saldo, ing, usc = _saldo(movimenti)
    ws["D1"], ws["E1"] = "SALDO", float(saldo)
    ws["D2"], ws["E2"] = "Tot ingressi", float(ing)
    ws["D3"], ws["E3"] = "Tot Uscite", float(usc)
    for r in (1, 2, 3):
        ws.cell(r, 4).font = _BOLD
        ws.cell(r, 5).number_format = "€ #,##0.00"

    for j, h in enumerate(_INTESTA, start=1):
        c = ws.cell(4, j, h)
        c.font = _BOLD
        c.fill = _HEADER_FILL
    for i, m in enumerate(sorted(movimenti, key=lambda x: x["data"]), start=5):
        importo = float(m["importo"]) * (1 if m["segno"] == "entrata" else -1)
        ws.cell(i, 1, m["data"]).number_format = "dd/mm/yyyy"
        ws.cell(i, 2, m.get("descrizione"))
        ws.cell(i, 3, m.get("n_fattura"))
        ws.cell(i, 4, "Entrata" if m["segno"] == "entrata" else "Uscita")
        ws.cell(i, 5, importo).number_format = "€ #,##0.00"
        ws.cell(i, 6, m.get("categoria"))
        ws.cell(
            i, 7, acronimi.get(str(m.get("iniziativa_id"))) or m.get("progetto_label")
        )
        ws.cell(i, 8, m.get("persona_contatto"))
        ws.cell(i, 9, m.get("note"))
    ws.column_dimensions["B"].width = 60
    for col in ("A", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 16


def _foglio_progetto(wb: Workbook, prog: dict, calendario: list[dict]) -> None:
    acr = prog.get("acronimo") or "PROGETTO"
    ws = wb.create_sheet(str(acr)[:31])
    info = [
        ("SEZIONE: INFO GENERALI", ""),
        ("Acronimo", acr),
        ("Nome esteso", prog.get("titolo") or ""),
        ("Ente finanziatore", prog.get("controparte") or ""),
        ("Costo complessivo", _num(prog.get("costo_complessivo"))),
        ("Finanziamento complessivo", _num(prog.get("finanziamento_complessivo"))),
        ("Data inizio", prog.get("data_inizio")),
        ("Data fine", prog.get("data_fine")),
    ]
    for r, (k, v) in enumerate(info, start=1):
        ws.cell(r, 1, k).font = _BOLD
        ws.cell(r, 2, v)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="EFEFEF")

    hr = 11
    ws.cell(hr, 1, "SEZIONE: CALENDARIO MOVIMENTI").font = _BOLD
    for j, h in enumerate(
        ["Descrizione", "Ingresso/Uscita", "Importo (€)", "Data attesa", "Completata"],
        start=1,
    ):
        c = ws.cell(hr + 1, j, h)
        c.font = _BOLD
        c.fill = _HEADER_FILL
    for i, m in enumerate(calendario, start=hr + 2):
        ws.cell(i, 1, m.get("descrizione"))
        ws.cell(i, 2, "Ingresso" if m["segno"] == "entrata" else "Uscita")
        ws.cell(i, 3, float(m["importo"])).number_format = "€ #,##0.00"
        ws.cell(i, 4, m.get("data_attesa"))
        ws.cell(i, 5, "Sì" if m.get("completata") else "")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16


def _num(v):
    return float(v) if v is not None else ""


def build_libro_cassa_xlsx(
    movimenti_per_anno: dict,
    acronimi_by_id: dict,
    progetti: list[dict] | None = None,
    calendari: dict | None = None,
) -> bytes:
    """Crea il workbook Libro Cassa. Ritorna i byte del file .xlsx.

    - `movimenti_per_anno`: {anno: [movimento dict]}.
    - `acronimi_by_id`: {iniziativa_id(str): acronimo} per la colonna Progetto.
    - `progetti` / `calendari`: opzionali, per i fogli per-progetto.
    """
    wb = Workbook()
    wb.remove(wb.active)
    for anno in sorted(movimenti_per_anno, key=lambda a: a or 0, reverse=True):
        _foglio_libro(wb, anno, movimenti_per_anno[anno], acronimi_by_id)
    for prog in progetti or []:
        cal = (calendari or {}).get(str(prog.get("id")), [])
        _foglio_progetto(wb, prog, cal)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def intestazione_titolo() -> Alignment:  # pragma: no cover - util di stile
    return Alignment(horizontal="center")
