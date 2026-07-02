"""Import del file contabile ANTECNICA ("Libro Cassa") esportato da Google Sheet.

Struttura del workbook:
- fogli «Libro Cassa <anno>»: righe 1-3 = SALDO/Tot ingressi/Tot Uscite;
  header sulla riga con «Data … Importo (€) … Progetto»; dati sotto.
  Colonne: Data, Descrizione Transazione, N. Fattura, Tipo Transazione
  (Entrata/Uscita), Importo (€) [con segno], Categoria, Progetto (acronimo),
  Persona/Contatto, Note.
- fogli per progetto (acronimo): INFO GENERALI (Acronimo, Nome esteso, Ente
  finanziatore, Costo complessivo, Finanziamento complessivo, Tot spese
  personale, Tot acquisti, Data inizio, Data fine) + CALENDARIO MOVIMENTI
  (Descrizione, Ingresso/Uscita, Importo, Data attesa, Completata).
- foglio «Spese periodiche»: Descrizione, Tipologia, Importo, Periodicità,
  Progetto, Dal, Al.
- foglio «Dati»: vocabolari (categorie uscite/entrate, acronimi, persone).

Le funzioni ritornano dati normalizzati (nessuna scrittura su DB).
"""

from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook

from src.domain.finanza import parse_data, parse_importo


def _norm(v: object) -> str:
    return str(v).strip().lower() if v is not None else ""


def _segno(v: object) -> str | None:
    s = _norm(v)
    if s in ("entrata", "ingresso", "e", "+"):
        return "entrata"
    if s in ("uscita", "u", "-"):
        return "uscita"
    return None


def _mappa_header_cassa(righe: list[list]) -> tuple[int, dict[str, int]] | None:
    """Trova la riga di intestazione del Libro Cassa e la mappa colonna->indice."""
    for i, riga in enumerate(righe[:15]):
        norm = [_norm(c) for c in riga]
        if "data" in norm and any("importo" in c for c in norm):
            mappa: dict[str, int] = {}
            for j, c in enumerate(norm):
                if c == "data":
                    mappa["data"] = j
                elif "descriz" in c:
                    mappa["descrizione"] = j
                elif "fattura" in c:
                    mappa["n_fattura"] = j
                elif "tipo" in c:
                    mappa["segno"] = j
                elif "importo" in c:
                    mappa["importo"] = j
                elif "categoria" in c:
                    mappa["categoria"] = j
                elif c == "progetto":
                    mappa["progetto"] = j
                elif "persona" in c:
                    mappa["persona"] = j
                elif c == "note":
                    mappa["note"] = j
            if "data" in mappa and "importo" in mappa:
                return i, mappa
    return None


def parse_libro_cassa(righe: list[list]) -> list[dict]:
    """Righe grezze (list of list) di un foglio Libro Cassa -> movimenti."""
    trovato = _mappa_header_cassa(righe)
    if trovato is None:
        return []
    header_idx, m = trovato
    out: list[dict] = []
    for riga in righe[header_idx + 1 :]:

        def cell(campo: str, riga=riga):
            j = m.get(campo)
            return riga[j] if j is not None and j < len(riga) else None

        d = parse_data(cell("data"))
        imp = parse_importo(cell("importo"))
        if d is None or imp is None:
            continue
        segno = _segno(cell("segno")) or ("entrata" if imp >= 0 else "uscita")
        prog = cell("progetto")
        pers = cell("persona")
        out.append(
            {
                "data": d,
                "importo": abs(imp),
                "segno": segno,
                "descrizione": (
                    str(cell("descrizione")).strip() if cell("descrizione") else None
                ),
                "categoria": (
                    str(cell("categoria")).strip() if cell("categoria") else None
                ),
                "n_fattura": (
                    str(cell("n_fattura")).strip() if cell("n_fattura") else None
                ),
                "persona_contatto": str(pers).strip() if pers else None,
                "note": str(cell("note")).strip() if cell("note") else None,
                "progetto_label": str(prog).strip() if prog else None,
            }
        )
    return out


_INFO_LABELS = {
    "acronimo": "acronimo",
    "nome esteso": "nome_esteso",
    "ente finanziatore": "ente_finanziatore",
    "costo complessivo": "costo_complessivo",
    "finanziamento complessivo": "finanziamento_complessivo",
    "data inizio": "data_inizio",
    "data fine": "data_fine",
}


def parse_progetto(righe: list[list]) -> dict:
    """Foglio di un progetto -> {info: {...}, calendario: [movimenti previsti]}."""
    info: dict = {}
    for riga in righe[:12]:
        if not riga:
            continue
        label = _norm(riga[0])
        val = riga[1] if len(riga) > 1 else None
        for prefisso, chiave in _INFO_LABELS.items():
            if label.startswith(prefisso):
                if chiave in ("data_inizio", "data_fine"):
                    info[chiave] = parse_data(val)
                elif chiave in ("costo_complessivo", "finanziamento_complessivo"):
                    imp = parse_importo(val)
                    info[chiave] = imp
                else:
                    info[chiave] = str(val).strip() if val else None
                break
        if label.startswith("tot spese personale"):
            info["tot_spese_personale"] = parse_importo(val)
        elif label.startswith("tot acquisti"):
            info["tot_acquisti"] = parse_importo(val)

    # calendario movimenti: header con "Descrizione" + "Ingresso/Uscita"
    cal_header = None
    for i, riga in enumerate(righe):
        norm = [_norm(c) for c in riga]
        if any("descrizione" in c for c in norm) and any(
            "ingresso" in c or c == "tipo" for c in norm
        ):
            cal_header = i
            break
    calendario: list[dict] = []
    if cal_header is not None:
        for riga in righe[cal_header + 1 :]:
            if not riga or all(c is None or str(c).strip() == "" for c in riga):
                continue
            descr = riga[0] if len(riga) > 0 else None
            segno = _segno(riga[1]) if len(riga) > 1 else None
            imp = parse_importo(riga[2]) if len(riga) > 2 else None
            data_attesa = parse_data(riga[3]) if len(riga) > 3 else None
            completata = (
                bool(riga[4]) if len(riga) > 4 and riga[4] not in (None, "") else False
            )
            if imp is None and not descr:
                continue
            calendario.append(
                {
                    "descrizione": str(descr).strip() if descr else None,
                    "segno": segno or "uscita",
                    "importo": abs(imp) if imp is not None else Decimal("0"),
                    "data_attesa": data_attesa,
                    "completata": completata,
                }
            )
    return {"info": info, "calendario": calendario}


def parse_spese_periodiche(righe: list[list]) -> list[dict]:
    """Foglio «Spese periodiche» -> righe normalizzate."""
    if not righe:
        return []
    header = [_norm(c) for c in righe[0]]

    def idx(*keys):
        for k in keys:
            for j, c in enumerate(header):
                if k in c:
                    return j
        return None

    j_desc = idx("descrizione")
    j_tip = idx("tipologia", "categoria")
    j_imp = idx("importo")
    j_per = idx("periodicit")
    j_prog = idx("progetto")
    j_dal = idx("dal")
    j_al = idx("al")
    out = []
    for riga in righe[1:]:
        if not riga or all(c is None or str(c).strip() == "" for c in riga):
            continue

        def g(j, riga=riga):
            return riga[j] if j is not None and j < len(riga) else None

        imp = parse_importo(g(j_imp))
        desc = g(j_desc)
        if not desc and imp is None:
            continue
        out.append(
            {
                "descrizione": str(desc).strip() if desc else None,
                "tipologia": str(g(j_tip)).strip() if g(j_tip) else None,
                "importo": abs(imp) if imp is not None else None,
                "periodicita": str(g(j_per)).strip() if g(j_per) else None,
                "progetto_label": str(g(j_prog)).strip() if g(j_prog) else None,
                "dal": parse_data(g(j_dal)),
                "al": parse_data(g(j_al)),
            }
        )
    return out


def _righe_foglio(ws) -> list[list]:
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append([_cell(v) for v in row])
    return out


def _cell(v):
    if isinstance(v, datetime):
        return v.date()
    return v


def leggi_workbook(contenuto: bytes) -> dict:
    """Legge tutto il workbook e ritorna le sezioni normalizzate.

    Ritorna: {
      "libri": {anno: [movimenti]},          # fogli Libro Cassa <anno>
      "progetti": {acronimo: {info, calendario}},
      "spese_periodiche": [...],
      "acronimi_noti": [str],                # dal foglio Dati
    }
    """
    wb = load_workbook(io.BytesIO(contenuto), data_only=True)
    libri: dict[int, list] = {}
    progetti: dict[str, dict] = {}
    spese: list = []
    acronimi: list[str] = []

    for name in wb.sheetnames:
        ws = wb[name]
        righe = _righe_foglio(ws)
        low = name.strip().lower()
        if low.startswith("libro cassa"):
            movimenti = parse_libro_cassa(righe)
            if not movimenti:
                continue
            # anno dal nome ("Libro Cassa 2026") o dai dati
            anno = None
            for tok in name.split():
                if tok.isdigit() and len(tok) == 4:
                    anno = int(tok)
            if anno is None and movimenti:
                anno = movimenti[0]["data"].year
            libri.setdefault(anno, []).extend(movimenti)
        elif low.startswith("spese periodiche"):
            spese = parse_spese_periodiche(righe)
        elif low == "dati":
            # colonna Acronimo progetto (col C = indice 2)
            for riga in righe[1:]:
                if len(riga) > 2 and riga[2]:
                    acronimi.append(str(riga[2]).strip())
        else:
            # foglio progetto: ha "SEZIONE: INFO GENERALI" in A1
            if righe and _norm(righe[0][0]).startswith("sezione"):
                parsed = parse_progetto(righe)
                acr = parsed["info"].get("acronimo") or name.strip()
                progetti[acr] = parsed
    return {
        "libri": libri,
        "progetti": progetti,
        "spese_periodiche": spese,
        "acronimi_noti": acronimi,
    }


def saldo(movimenti: list[dict]) -> Decimal:
    tot = Decimal("0")
    for m in movimenti:
        tot += m["importo"] if m["segno"] == "entrata" else -m["importo"]
    return tot


def riepilogo(sezioni: dict) -> dict:
    return {
        "anni": sorted(a for a in sezioni["libri"] if a),
        "n_movimenti": sum(len(v) for v in sezioni["libri"].values()),
        "n_progetti": len(sezioni["progetti"]),
        "n_spese_periodiche": len(sezioni["spese_periodiche"]),
    }
